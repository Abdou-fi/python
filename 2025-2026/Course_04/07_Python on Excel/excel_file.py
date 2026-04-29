## penpyxl - automate Excel Tasks

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws["A1"] = "Hello"
ws["B1"] = "World"
wb.save("demo.xlsx")
print("Excel file created")